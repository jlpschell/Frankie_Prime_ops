#!/usr/bin/env python3
"""Process new Outscraper exports into leads_v2 format."""
import openpyxl
import csv
import os

INBOUND = "/home/plotting1/.openclaw/media/inbound"
LEADS_V2 = "/home/plotting1/frankie-bot/workspace/leads_v2"

FILES = {
    "88e8a0df-a4c8-4d16-93ca-3b91271fb5ef.xlsx": "garage_door",
    "2c9b5e9d-590e-4b0d-9215-1823070b39b9.xlsx": "concrete",
    "595cc4d6-c40e-42d9-b738-ca5a67874f63.xlsx": "fence",
    "fbfd0e59-70b8-47b9-8445-008f7eb2dea9.xlsx": "landscaping",
    "140e416d-11d5-4860-b0f9-a77304d57c54.xlsx": "electricians",
    "09ba3533-381d-4d0c-8a42-2bd7d1d985b1.xlsx": "plumbers",
    "2aa411e2-c8d5-4e9e-8eca-eb29ac91dffe.xlsx": "roofers",
    "a88e7db0-93ad-4e15-ab4a-32f58f3512c0.xlsx": "hvac",
    "dcc99550-6c02-4ee5-b838-e83122bcc9ec.xlsx": "general_contractors",
    "8fb48399-2212-48f4-9a2f-79d3ae2d8811.xlsx": "medspas",
    "1e0d3f07-8166-4b2e-bbcd-95ade7a91099.xlsx": "attorneys",
}

KEY_COLS = [
    'name', 'name_for_emails', 'subtypes', 'category', 'phone',
    'phone.whitepages_phones.lookup_type', 'phone.whitepages_phones.name',
    'website', 'address', 'street', 'city', 'state_code', 'postal_code',
    'rating', 'reviews', 'verified',
    'email_1', 'email_1.emails_validator.status', 'email_1_full_name', 'email_1_first_name', 'email_1_last_name', 'email_1_title',
    'email_2', 'email_2.emails_validator.status',
    'phone_1', 'phone_1.whitepages_phones.lookup_type', 'phone_1.whitepages_phones.name', 'phone_1.whitepages_phones.phone_type',
    'phone_2', 'phone_2.whitepages_phones.lookup_type',
    'facebook', 'instagram', 'linkedin',
    'company_insights.employees', 'company_insights.revenue', 'company_insights.industry',
]

def normalize_phone(p):
    if not p: return None
    s = str(p).strip().replace(' ','').replace('-','').replace('(','').replace(')','').replace('+','')
    # Remove leading 1 check
    if len(s) == 10: s = '1' + s
    if len(s) == 11 and s.startswith('1'): return '+' + s
    return None

total_all = 0
for fname, niche in FILES.items():
    fpath = os.path.join(INBOUND, fname)
    if not os.path.exists(fpath):
        print(f"SKIP {niche}: file not found")
        continue
    
    wb = openpyxl.load_workbook(fpath, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    
    col_map = {}
    for kc in KEY_COLS:
        if kc in headers:
            col_map[kc] = headers.index(kc)
    
    outdir = os.path.join(LEADS_V2, niche)
    os.makedirs(outdir, exist_ok=True)
    outfile = os.path.join(outdir, f"{niche}_master_v2.csv")
    
    rows_written = 0
    phones_for_twilio = []
    
    with open(outfile, 'w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=KEY_COLS + ['phone_normalized', 'niche'])
        writer.writeheader()
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            record = {}
            for kc in KEY_COLS:
                idx = col_map.get(kc)
                record[kc] = row[idx] if idx is not None else None
            
            norm = normalize_phone(record.get('phone'))
            if not norm:
                norm = normalize_phone(record.get('phone_1'))
            record['phone_normalized'] = norm
            record['niche'] = niche
            
            writer.writerow(record)
            rows_written += 1
            
            if norm:
                phones_for_twilio.append(norm)
    
    wb.close()
    
    phones_file = os.path.join(outdir, f"{niche}_phones_to_verify.txt")
    with open(phones_file, 'w') as f:
        for p in phones_for_twilio:
            f.write(p + '\n')
    
    print(f"✅ {niche}: {rows_written} leads, {len(phones_for_twilio)} phones to verify → {outfile}")
    total_all += rows_written

print(f"\n🎯 TOTAL: {total_all} leads processed across {len(FILES)} niches")
