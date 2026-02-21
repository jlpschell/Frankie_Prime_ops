#!/usr/bin/env python3
"""Extract mobile-verified leads from Outscraper exports and prep for GHL upload."""
import openpyxl
import csv
import os
import json
from datetime import datetime

INBOUND = "/home/plotting1/.openclaw/media/inbound"
OUTPUT_DIR = "/home/plotting1/.openclaw/workspace/ghl_upload"
os.makedirs(OUTPUT_DIR, exist_ok=True)

TODAY = datetime.now().strftime("%m-%d-%y")
LOAD_DATE = datetime.now().strftime("%Y-%m-%d")

# Files with carrier_type built in
FILES_WITH_CARRIER = {
    "2c9b5e9d-590e-4b0d-9215-1823070b39b9.xlsx": ("concrete", "chris"),
    "595cc4d6-c40e-42d9-b738-ca5a67874f63.xlsx": ("fence", "eric"),
    "fbfd0e59-70b8-47b9-8445-008f7eb2dea9.xlsx": ("landscaping", "chris"),
    "140e416d-11d5-4860-b0f9-a77304d57c54.xlsx": ("electricians", "eric"),
    "09ba3533-381d-4d0c-8a42-2bd7d1d985b1.xlsx": ("plumbers", "chris"),
    "2aa411e2-c8d5-4e9e-8eca-eb29ac91dffe.xlsx": ("roofers", "eric"),
    "a88e7db0-93ad-4e15-ab4a-32f58f3512c0.xlsx": ("hvac", "chris"),
    "dcc99550-6c02-4ee5-b838-e83122bcc9ec.xlsx": ("general_contractors", "eric"),
    "6759c7f8-6661-4e16-bd8f-dfda50ccda9a.xlsx": ("pest_control", "email_only"),
}

# Garage door has no carrier type - will need Twilio
FILES_NO_CARRIER = {
    "88e8a0df-a4c8-4d16-93ca-3b91271fb5ef.xlsx": ("garage_door", "eric"),
}

def normalize_phone(p):
    if not p: return None
    s = str(p).strip().replace(' ','').replace('-','').replace('(','').replace(')','').replace('+','')
    if len(s) == 10: s = '1' + s
    if len(s) == 11 and s.startswith('1'): return '+' + s
    return None

all_mobile = []
niche_counts = {}

for fname, (niche, voice) in FILES_WITH_CARRIER.items():
    if voice == "email_only":
        continue
    fpath = os.path.join(INBOUND, fname)
    wb = openpyxl.load_workbook(fpath, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    
    # Find carrier type column
    carrier_idx = headers.index('phone.phones_enricher.carrier_type')
    
    # Find key columns
    def gi(name):
        return headers.index(name) if name in headers else None
    
    name_i = gi('name')
    phone_i = gi('phone')
    email_i = gi('email_1')
    email_status_i = gi('email_1.emails_validator.status')
    website_i = gi('website')
    city_i = gi('city')
    state_i = gi('state_code') or gi('state')
    zip_i = gi('postal_code')
    address_i = gi('address')
    contact_name_i = gi('email_1_full_name')
    first_i = gi('email_1_first_name')
    last_i = gi('email_1_last_name')
    rating_i = gi('rating')
    reviews_i = gi('reviews')
    
    mobile_count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        carrier = str(row[carrier_idx]).lower() if row[carrier_idx] else ''
        if carrier != 'mobile':
            continue
        
        phone_norm = normalize_phone(row[phone_i] if phone_i is not None else None)
        if not phone_norm:
            continue
        
        # Get validated email only
        email = row[email_i] if email_i is not None else None
        email_status = str(row[email_status_i]).upper() if email_status_i is not None and row[email_status_i] else ''
        if email_status not in ('RECEIVING', 'VALID'):
            email = None
        
        lead = {
            'business_name': row[name_i] if name_i is not None else '',
            'phone': phone_norm,
            'email': email or '',
            'website': row[website_i] if website_i is not None else '',
            'address': row[address_i] if address_i is not None else '',
            'city': row[city_i] if city_i is not None else '',
            'state': row[state_i] if state_i is not None else '',
            'zip': row[zip_i] if zip_i is not None else '',
            'first_name': row[first_i] if first_i is not None else '',
            'last_name': row[last_i] if last_i is not None else '',
            'rating': row[rating_i] if rating_i is not None else '',
            'reviews': row[reviews_i] if reviews_i is not None else '',
            'niche': niche,
            'voice': voice,
            'carrier_type': 'mobile',
            'load_date': LOAD_DATE,
            'tags': f"{niche}_{TODAY},mobile-verified,cold-outreach,voice-{voice},load-{TODAY}",
        }
        all_mobile.append(lead)
        mobile_count += 1
    
    niche_counts[niche] = mobile_count
    wb.close()
    print(f"✅ {niche}: {mobile_count} mobile leads")

# Write master CSV for GHL import
master_file = os.path.join(OUTPUT_DIR, f"ghl_mobile_master_{TODAY}.csv")
ghl_fields = ['business_name', 'phone', 'email', 'website', 'address', 'city', 'state', 'zip',
              'first_name', 'last_name', 'rating', 'reviews', 'niche', 'voice', 'carrier_type', 
              'load_date', 'tags']

with open(master_file, 'w', newline='') as f:
    writer = csv.DictWriter(f, fieldnames=ghl_fields)
    writer.writeheader()
    for lead in all_mobile:
        writer.writerow(lead)

# Also write per-niche files
for niche in niche_counts:
    niche_leads = [l for l in all_mobile if l['niche'] == niche]
    niche_file = os.path.join(OUTPUT_DIR, f"ghl_{niche}_{TODAY}.csv")
    with open(niche_file, 'w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=ghl_fields)
        writer.writeheader()
        for lead in niche_leads:
            writer.writerow(lead)

# Write JSON for API upload
json_file = os.path.join(OUTPUT_DIR, f"ghl_mobile_master_{TODAY}.json")
with open(json_file, 'w') as f:
    json.dump(all_mobile, f, indent=2)

print(f"\n🎯 TOTAL MOBILE LEADS: {len(all_mobile)}")
print(f"📁 Master CSV: {master_file}")
print(f"📁 Master JSON: {json_file}")
print(f"\nPer-niche breakdown:")
for n, c in sorted(niche_counts.items(), key=lambda x: -x[1]):
    print(f"  {n}: {c}")
